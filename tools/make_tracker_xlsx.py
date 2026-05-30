#!/usr/bin/env python3
"""Build an Excel + CSV tracker from reports/, with a per-job Word resume.

For each evaluated role this:
  - parses the report (Machine Summary + header) for company/role/score/URL/PDF,
  - ensures a per-job .docx resume exists at
      output/resume-{num:03d}-{company-slug}-{role-slug}.docx
    (generated from cv.md; this name is the row's primary key — it ties the
    job title to its resume file and is easy to find in output/),
  - writes data/applications.xlsx and data/applications.csv with Job URL,
    Resume (.docx) and Resume (PDF) columns adjacent for quick access.

Re-runnable and idempotent (existing .docx are not regenerated). Zero LLM cost.
"""
import os, re, glob, sys, shutil, subprocess, html as _html

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPORTS = os.path.join(ROOT, "reports")
OUT = os.path.join(ROOT, "data", "applications.xlsx")
CV = os.path.join(ROOT, "cv.md")
BASE_PDF = os.path.join(ROOT, "output", "resume-base.pdf")
STATUS_FILE = os.path.join(ROOT, "data", "applied-status.tsv")


def load_status():
    """num(str) -> status, from data/applied-status.tsv (last write wins)."""
    m = {}
    if os.path.exists(STATUS_FILE):
        for line in open(STATUS_FILE, encoding="utf-8"):
            p = line.rstrip("\n").split("\t")
            if len(p) >= 2 and p[0].isdigit():
                m[p[0]] = p[1]
    return m

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from generate_docx import build as build_docx  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

HEADER_RE = lambda k: re.compile(r"^\*\*%s:\*\*\s*(.+?)\s*$" % k, re.M)
H1_RE = re.compile(r"^#\s*Evaluation:\s*(.+?)\s*$", re.M)
MS_RE = re.compile(r"```yaml\s*(.*?)```", re.S)


def slug(s, maxlen=40):
    s = re.sub(r"[^a-z0-9]+", "-", (s or "").lower()).strip("-")
    return s[:maxlen].strip("-") or "na"


def ms_value(ms, key):
    m = re.search(r'^%s:\s*"?(.*?)"?\s*$' % re.escape(key), ms, re.M)
    return m.group(1).strip() if m else ""


def ms_first_list_item(ms, key):
    m = re.search(r'^%s:\s*\n((?:\s*-\s*.+\n?)+)' % re.escape(key), ms, re.M)
    if not m:
        return ""
    return re.sub(r'^\s*-\s*"?(.*?)"?\s*$', r"\1", m.group(1).splitlines()[0])


def rel(path):
    path = path.strip().lstrip("/")
    if path.startswith("career-ops/"):
        path = path[len("career-ops/"):]
    return path


def parse_report(fp):
    txt = open(fp, encoding="utf-8").read()
    base = os.path.basename(fp)
    num = int(base.split("-", 1)[0])
    h1 = H1_RE.search(txt)
    company, role = "", ""
    if h1:
        parts = re.split(r"\s+[—-]\s+", h1.group(1), maxsplit=1)
        company = parts[0].strip()
        role = parts[1].strip() if len(parts) > 1 else ""

    def hdr(k):
        m = HEADER_RE(k).search(txt)
        return m.group(1).strip() if m else ""

    ms_m = MS_RE.search(txt)
    ms = ms_m.group(1) if ms_m else ""
    if ms:
        company = ms_value(ms, "company") or company
        role = ms_value(ms, "role") or role

    score = (hdr("Score") or (ms_value(ms, "score") + "/5")).replace("/5/5", "/5")
    return {
        "num": num,
        "date": hdr("Date"),
        "company": company,
        "role": role,
        "score": score,
        "decision": ms_value(ms, "final_decision"),
        "legitimacy": hdr("Legitimacy"),
        "url": hdr("URL"),
        "report": os.path.relpath(fp, ROOT),
        "pdf": rel(hdr("PDF")),
        "note": ms_value(ms, "next_action") or ms_first_list_item(ms, "hard_stops"),
    }


def job_stem(row):
    return "resume-%03d-%s-%s" % (row["num"], slug(row["company"], 24), slug(row["role"]))


def ensure_base_pdf():
    """Render the base resume PDF once (cv.md -> HTML -> generate-pdf.mjs)."""
    if os.path.exists(BASE_PDF):
        return True
    try:
        html = "/tmp/resume-base.html"
        subprocess.run([sys.executable, os.path.join(os.path.dirname(__file__), "md_to_html.py"),
                        CV, html], check=True, cwd=ROOT)
        subprocess.run(["node", "generate-pdf.mjs", html, BASE_PDF], check=True, cwd=ROOT)
        return os.path.exists(BASE_PDF)
    except Exception as e:  # noqa: BLE001
        print(f"  warn: base PDF render failed: {e}", file=sys.stderr)
        return False


def ensure_job_docx(row):
    """Per-job resume path = the primary key. Generate from cv.md if missing."""
    relpath = os.path.join("output", job_stem(row) + ".docx")
    abspath = os.path.join(ROOT, relpath)
    if not os.path.exists(abspath) and os.path.exists(CV):
        try:
            build_docx(CV, abspath)
        except Exception as e:  # noqa: BLE001
            print(f"  warn: docx for #{row['num']} failed: {e}", file=sys.stderr)
            return ""
    return relpath


def ensure_job_pdf(row, base_ok):
    """Per-job PDF mirrors the .docx name; copy from the base render if missing."""
    relpath = os.path.join("output", job_stem(row) + ".pdf")
    abspath = os.path.join(ROOT, relpath)
    if not os.path.exists(abspath) and base_ok:
        try:
            shutil.copyfile(BASE_PDF, abspath)
        except Exception as e:  # noqa: BLE001
            print(f"  warn: pdf for #{row['num']} failed: {e}", file=sys.stderr)
            return ""
    return relpath if os.path.exists(abspath) else ""


def main():
    rows = [parse_report(fp) for fp in glob.glob(os.path.join(REPORTS, "*.md"))]
    rows.sort(key=lambda r: r["num"])
    status_map = load_status()
    base_ok = ensure_base_pdf()
    made = 0
    for r in rows:
        existed = os.path.exists(os.path.join(ROOT, "output", job_stem(r) + ".docx"))
        r["docx"] = ensure_job_docx(r)
        r["pdf"] = ensure_job_pdf(r, base_ok)
        if r["docx"] and not existed:
            made += 1

    cols = ["#", "Company", "Role", "Score", "Decision", "Job URL",
            "Resume (.docx)", "Resume (PDF)", "Report (JD details)",
            "Legitimacy", "Date", "Status", "Notes"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(cols)
    head_fill = PatternFill("solid", fgColor="1F2A44")
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "B2"

    def fileurl(relpath):
        return "file://" + os.path.join(ROOT, relpath)

    def link(r, c, target, text, is_url=False):
        cell = ws.cell(r, c, text or "")
        if target and (is_url or os.path.exists(os.path.join(ROOT, target))):
            cell.hyperlink = target if is_url else fileurl(target)
            cell.font = Font(color="0563C1", underline="single")

    for i, row in enumerate(rows, start=2):
        ws.cell(i, 1, row["num"])
        ws.cell(i, 2, row["company"])
        ws.cell(i, 3, row["role"])
        ws.cell(i, 4, row["score"])
        ws.cell(i, 5, row["decision"])
        link(i, 6, row["url"], row["url"], is_url=True)
        link(i, 7, row["docx"], os.path.basename(row["docx"]) if row["docx"] else "")
        link(i, 8, row["pdf"], "PDF")
        link(i, 9, row["report"], "report")
        ws.cell(i, 10, row["legitimacy"])
        ws.cell(i, 11, row["date"])
        ws.cell(i, 12, status_map.get(str(row["num"]), "Evaluated"))
        ws.cell(i, 13, row["note"])

    widths = [5, 18, 34, 8, 14, 46, 44, 8, 9, 18, 11, 11, 60]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)

    # CSV: emit =HYPERLINK() formulas so links are clickable in Excel/Numbers.
    # Web URL works everywhere; file:// links open only on this Mac (not in
    # Google Sheets, where local files are blocked by design).
    import csv

    def hl(target, label, is_url=False):
        if not target:
            return ""
        if not is_url and not os.path.exists(os.path.join(ROOT, target)):
            return ""
        url = target if is_url else "file://" + os.path.join(ROOT, target)
        return '=HYPERLINK("%s","%s")' % (url.replace('"', "%22"), label.replace('"', "'"))

    csv_path = OUT[:-5] + ".csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(cols)
        for row in rows:
            w.writerow([
                row["num"], row["company"], row["role"], row["score"], row["decision"],
                hl(row["url"], "open", is_url=True),
                hl(row["docx"], os.path.basename(row["docx"]) if row["docx"] else ""),
                hl(row["pdf"], "PDF"),
                hl(row["report"], "report"),
                row["legitimacy"], row["date"],
                status_map.get(str(row["num"]), "Evaluated"), row["note"],
            ])

    # HTML tracker: opens in a browser where file:// links to local PDFs/docx
    # actually work (unlike sandboxed Excel / Numbers). This is the click-through view.
    def a(target, label, is_url=False):
        if not target:
            return ""
        if not is_url and not os.path.exists(os.path.join(ROOT, target)):
            return ""
        href = target if is_url else "file://" + os.path.join(ROOT, target)
        return f'<a href="{href}"{" target=_blank" if is_url else ""}>{_html.escape(label)}</a>'

    trs = []
    for row in rows:
        trs.append(
            "<tr>"
            f"<td>{row['num']}</td><td>{_html.escape(row['company'])}</td>"
            f"<td>{_html.escape(row['role'])}</td><td class=sc>{_html.escape(row['score'])}</td>"
            f"<td>{_html.escape(row['decision'])}</td>"
            f"<td>{_html.escape(status_map.get(str(row['num']), 'Evaluated'))}</td>"
            f"<td>{a(row['url'],'🔗 job', is_url=True)}</td>"
            f"<td>{a(row['docx'],'📝 docx')}</td>"
            f"<td>{a(row['pdf'],'📄 pdf')}</td>"
            f"<td>{a(row['report'],'report')}</td>"
            f"<td>{_html.escape(row['note'])}</td></tr>"
        )
    out_dir_url = "file://" + os.path.join(ROOT, "output")
    html_doc = (
        "<!doctype html><meta charset=utf-8><title>career-ops tracker</title>"
        "<style>body{font:14px -apple-system,Segoe UI,sans-serif;margin:24px}"
        "h1{font-size:20px}table{border-collapse:collapse;width:100%}"
        "th,td{border-bottom:1px solid #ddd;padding:6px 8px;text-align:left;vertical-align:top}"
        "th{position:sticky;top:0;background:#1F2A44;color:#fff}tr:nth-child(even){background:#f6f7f9}"
        "td.sc{font-weight:700}a{color:#0563C1;text-decoration:none}a:hover{text-decoration:underline}"
        ".bar{margin:8px 0 16px}.bar a{background:#1F2A44;color:#fff;padding:6px 10px;border-radius:6px}</style>"
        f"<h1>career-ops — {len(rows)} roles</h1>"
        f'<div class=bar><a href="{out_dir_url}">📁 Open resumes folder</a></div>'
        "<table><tr><th>#</th><th>Company</th><th>Role</th><th>Score</th><th>Decision</th>"
        "<th>Status</th><th>Job</th><th>Resume</th><th>PDF</th><th>Report</th><th>Notes</th></tr>"
        + "".join(trs) + "</table>"
    )
    html_path = OUT[:-5] + ".html"
    open(html_path, "w", encoding="utf-8").write(html_doc)

    print(f"wrote {OUT}, {csv_path}, {html_path} — {len(rows)} rows; generated {made} new per-job .docx")


if __name__ == "__main__":
    main()
